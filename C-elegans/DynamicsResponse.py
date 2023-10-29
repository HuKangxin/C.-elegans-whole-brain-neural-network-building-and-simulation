# Created by Kangxin Hu at Beihang University, on Oct.2023.

from __future__ import division
from decoding import decode1
from decoding import decode2
from decoding import limit
import socket
import string
sock1 = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)  # ipv4,udp
sock2 = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
sock1.bind(('192.168.31.165', 54377))  # UDP服务器端口和IP绑定
  # 等待matlab发送请求，这样就能获取matlab client的ip和端口号
import csv

  # 将数据转为bytes发送给matlab的client

from CreatNeurons import *
from GeneToG import *
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
import time
import random

h.load_file('stdrun.hoc')
h.nrn_load_dll('/mod/nrnmech.dll')

plt.ion()
#################################################### Build Model #######################################################
Neu_workbook = openpyxl.load_workbook("neuronlist.xlsx")
Neu_worksheet = Neu_workbook["Sheet1"]
neuron_list = []
for num in range(2,304):
     neuron_list.append(Neu_worksheet.cell(row=num,column=1).value)
 # a list that contains name of 302 elements

Gene_workbook = openpyxl.load_workbook("/data/GenesExpressingData.xlsx")
Gene_worksheet = Gene_workbook["Sheet2"]

Con_workbook = openpyxl.load_workbook("/data/Table_Export2.xlsx")
Con_worksheet = Con_workbook["Sheet1"]

muscle_DB = openpyxl.load_workbook("/data/DB.xlsx")
DB_worksheet = muscle_DB["Sheet1"]

muscle_VB = openpyxl.load_workbook("/data/VB.xlsx")
VB_worksheet = muscle_VB["Sheet1"]

h.celsius = 30
shapePara = 7  # um # length and diameter of soma
RPara = 10      # axial resistance
CPara = 3.1     # membrane capacitance
Para = np.zeros(23)     # a list that contains a series of values of neuronal signals

neurons = []  # a list that contains the built model of 302 elements, which are neurons of C.elegans
for i in range(1, 303):
    index = i+1

    shl1 = Gene_worksheet.cell(row=index, column=5).value
    shk1 = Gene_worksheet.cell(row=index, column=6).value
    kvs1 = Gene_worksheet.cell(row=index, column=8).value
    kqt3 = Gene_worksheet.cell(row=index, column=9).value
    egl2 = Gene_worksheet.cell(row=index, column=6).value
    egl36 = Gene_worksheet.cell(row=index, column=15).value
    irk = Gene_worksheet.cell(row=index, column=14).value + Gene_worksheet.cell(row=index, column=13).value + Gene_worksheet.cell(row=index, column=12).value
    egl19 = Gene_worksheet.cell(row=index, column=17).value
    unc2 = Gene_worksheet.cell(row=index, column=2).value
    cca1 = Gene_worksheet.cell(row=index, column=18).value
    slo1 = Gene_worksheet.cell(row=index, column=4).value
    slo2 = Gene_worksheet.cell(row=index, column=3).value
    kcnl = Gene_worksheet.cell(row=index, column=11).value
    nca = Gene_worksheet.cell(row=index, column=7).value
    trp4 = Gene_worksheet.cell(row=index, column=19).value
    itr1 = Gene_worksheet.cell(row=index, column=20).value
    osm9 = Gene_worksheet.cell(row=index, column=21).value
    mec10 = Gene_worksheet.cell(row=index, column=22).value

    Para[0], Para[1], Para[2], Para[3], Para[4], Para[5], Para[6], Para[7], Para[8], Para[9], Para[10], Para[12], Para[14], Para[15], Para[19], Para[20], Para[21], \
    Para[22] = GeneToG(shl1, shk1, kvs1, kqt3, egl2, egl36, irk, egl19, unc2, cca1, slo1, slo2, kcnl, nca, itr1, mec10, osm9, trp4)

    Para[11] = Para[10]
    Para[13] = Para[12]
    Para[16] = 0.289   # gene 索引无 leak, 自己定的
    Para[17] = -84
    Para[18] = 55

    # print(Con_worksheet.cell(row=i+4, column=3).value)
    neurons.append(Cell(shapePara, RPara, CPara, Para))
#################################################### Connectom #########################################################
print(neuron_list)
post_neurons = {} # a dictionary that contains the post neurons corresponding to the pre_neurons
polarity = {} # a dictionary that contains the polarity information corresponding to the pre_neurons
weight = {} # a dictionary that contains the weight value corresponding to the pre_neurons
m=2
for i in range(0,302):
    post = []
    weight_ = []
    polarity_ = []
    for j in range(m,3175):
        if Con_worksheet.cell(row=j,column=1).value == neuron_list[i]:
            post.append(Con_worksheet.cell(row=j,column=2).value)
            weight_.append(Con_worksheet.cell(row=j, column=3).value)
            polarity_.append(Con_worksheet.cell(row=j, column=4).value)
        else:
            m=j
            break
    post_neurons[i] = post
    weight[i] = weight_
    polarity[i] = polarity_

syn = [] # a list that contains the polarity of post neurons
nc = [] # a list that contains the information of connection
thres = -30
Delay = 0
synweight = 1
for i in range(0, 302):
    for j in range(0,len(post_neurons[i])):
        neuron = post_neurons[i][j]
        index = neuron_list.index(neuron) # the sequence of the post neurons
        if weight[i][j] != 0:
            if polarity[i][j] == '+':
                Hebbsyn = h.Exc2Syn(neurons[index].soma(0.5))
                #h.setpointer(neurons[i].soma(0.5)._ref_v, 'Vpre', Hebbsyn)
                pre_post = h.NetCon(neurons[i].soma(0.5)._ref_v, Hebbsyn, sec=neurons[i].soma)
                synweight = weight[i][j]
                pre_post.weight[0] = synweight*0.5  # Con_worksheet.cell(row=irow, column=jcol).value
                pre_post.delay = Delay
                pre_post.threshold = thres
                syn.append(Hebbsyn)
                nc.append(pre_post)
            if polarity[i][j] == '-':
                Hebbsyn = h.InhSyn(neurons[index].soma(0.5))
                #h.setpointer(neurons[index].soma(0.5)._ref_v, 'Vpre', Hebbsyn)
                pre_post = h.NetCon(neurons[i].soma(0.5)._ref_v, Hebbsyn, sec=neurons[i].soma)
                synweight = weight[i][j]
                pre_post.weight[0] = synweight*0.5  # Con_worksheet.cell(row=irow, column=jcol).value
                pre_post.delay = Delay
                pre_post.threshold = thres
                syn.append(Hebbsyn)
                nc.append(pre_post)
#Connect the relationship between muscles and motor neurons
DB = {}
nm = 59
for x in range(0,6):
    motor_con = []
    for i in range(nm,75):
        if DB_worksheet.cell(row=i, column=2).value == DB_worksheet.cell(row=i+1, column=2).value :
            motor_neuron = DB_worksheet.cell(row=i, column=1).value
            j = neuron_list.index(motor_neuron)
            w = DB_worksheet.cell(row=i, column=4).value
            motor_con.append([j,w])
            nm += 1
        else :
            motor_neuron = DB_worksheet.cell(row=i, column=1).value
            j = neuron_list.index(motor_neuron)
            w = DB_worksheet.cell(row=i, column=4).value
            motor_con.append([j,w])
            DB[x] = motor_con
            nm += 1
            break

VB = {}
nm = 55
for x in range(0,6):
    motor_con = []
    for i in range(nm,75):
        if VB_worksheet.cell(row=i, column=2).value == VB_worksheet.cell(row=i+1, column=2).value :
            motor_neuron = VB_worksheet.cell(row=i, column=1).value
            j = neuron_list.index(motor_neuron)
            w = VB_worksheet.cell(row=i, column=4).value
            motor_con.append([j,w])
            nm += 1
        else :
            motor_neuron = VB_worksheet.cell(row=i, column=1).value
            j = neuron_list.index(motor_neuron)
            w = VB_worksheet.cell(row=i, column=4).value
            motor_con.append([j,w])
            VB[x] = motor_con
            nm += 1
            break
print(VB)
print('model built')

#################################################### Set up experiment #################################################

is_receiving=True
while(is_receiving):
    print('waiting...')
    buf, addr = sock1.recvfrom(40960)
    district=buf
    num=float(district)
    print(num)
    stim = h.IClamp(neurons[153].soma(0.5))  # add a current clamp at the middle of the soma
    stim.dur = 20  # ms
    stim.delay = 3
    stim.amp = 51.4+num*10
#################################################### Record Simulation #################################################
    neurons_v = []
    for k in range(1, 303):
        neurons_v.append(h.Vector())
        neurons_v[k-1].record(neurons[k-1].soma(0.5)._ref_v)
#################################################### Run Simulation ####################################################
    t = h.Vector()
    t.record(h._ref_t)  # record time
    h.v_init = -65  # set starting voltage
    h.tstop = 30  # set simulation time
    h.dt = 0.01
    h.run()  # run simulation
############################################### Fluorescence Intensity #################################################
    CalciumConcen = np.zeros((200001, 6))
    FluoIntensity = np.zeros((200001, 6))
    fmin = 0.81
    fmax = 9.7
    hFI = 2
    kd = 8.3
    k = 1
    V0 = 1
############################################### Output to muscles #################################################
    order1 = decode2(neurons_v[208])
    output=[order1]
    s=str(output)
    sock1.sendto(bytes(s, encoding="utf8"), addr)
    print(s)
    endtime=time.time()

#################################################### Plot Results ######################################################
'''
plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[2],  label='ADEL')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()
plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[195],  label='RIBL')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()
plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[76],  label='AWCL')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()
plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[196],  label='RIBR')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[213],  label='RMDDL')
plt.plot(t, neurons_v[214],  label='RMDDR')
plt.plot(t, neurons_v[215],  label='RMDL')
plt.plot(t, neurons_v[216],  label='RMDR')
plt.plot(t, neurons_v[217],  label='RMDVL')
plt.plot(t, neurons_v[218],  label='RMDVR')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[88],  label='DA1')
plt.plot(t, neurons_v[89],  label='DA2')
plt.plot(t, neurons_v[90],  label='DA3')
plt.plot(t, neurons_v[91],  label='DA4')
plt.plot(t, neurons_v[92],  label='DA5')
plt.plot(t, neurons_v[93],  label='DA6')
plt.plot(t, neurons_v[94],  label='DA7')
plt.plot(t, neurons_v[95],  label='DA8')
plt.plot(t, neurons_v[96],  label='DA9')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[97],  label='DB1')
plt.plot(t, neurons_v[98],  label='DB2')
plt.plot(t, neurons_v[99],  label='DB3')
plt.plot(t, neurons_v[100],  label='DB4')
plt.plot(t, neurons_v[103],  label='DB7')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[101],  label='DB5')
plt.plot(t, neurons_v[102],  label='DB6')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[266],  label='VA1')
plt.plot(t, neurons_v[267],  label='VA2')
plt.plot(t, neurons_v[268],  label='VA3')
plt.plot(t, neurons_v[269],  label='VA4')
plt.plot(t, neurons_v[270],  label='VA5')
plt.plot(t, neurons_v[271],  label='VA6')
plt.plot(t, neurons_v[272],  label='VA7')
plt.plot(t, neurons_v[273],  label='VA8')
plt.plot(t, neurons_v[275],  label='VA10')
plt.plot(t, neurons_v[276],  label='VA11')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[274],  label='VA9')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[277],  label='VA12')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[278],  label='VB1')
plt.plot(t, neurons_v[279],  label='VB2')
plt.plot(t, neurons_v[280],  label='VB3')
plt.plot(t, neurons_v[281],  label='VB4')
plt.plot(t, neurons_v[282],  label='VB5')
plt.plot(t, neurons_v[283],  label='VB6')
plt.plot(t, neurons_v[284],  label='VB7')
plt.plot(t, neurons_v[285],  label='VB8')
plt.plot(t, neurons_v[286],  label='VB9')
plt.plot(t, neurons_v[287],  label='VB10')
plt.plot(t, neurons_v[288],  label='VB11')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[246],  label='SMBDL')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[247],  label='SMBDR')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[248],  label='SMBVL')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

plt.figure(figsize=(8, 5))
plt.plot(t, neurons_v[249],  label='SMBVR')
plt.xlim(0, h.tstop)
plt.xlabel('Time (ms)', fontsize=15)
plt.ylabel('V (mV)', fontsize=15)
plt.legend(fontsize=14, frameon=False)
plt.tight_layout()

# ## CalciumConcen

plt.ioff()
'''